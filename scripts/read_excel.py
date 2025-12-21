#!/usr/bin/env python3
"""
Script temporário para ler a planilha Excel e mostrar sua estrutura
"""
import sys
import json

try:
    import openpyxl
    
    def read_excel(file_path):
        """Lê o arquivo Excel e retorna a estrutura"""
        wb = openpyxl.load_workbook(file_path)
        result = {
            "sheets": [],
            "structure": {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            
            # Ler até 20 linhas para entender a estrutura
            for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if idx > 20:  # Limitar a 20 linhas por sheet
                    break
                if any(cell is not None for cell in row):  # Ignorar linhas vazias
                    rows.append([str(cell) if cell is not None else '' for cell in row])
            
            result["sheets"].append({
                "name": sheet_name,
                "rows": rows,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            })
            
            # Primeira linha geralmente é o cabeçalho
            if rows:
                result["structure"][sheet_name] = {
                    "headers": rows[0] if len(rows) > 0 else [],
                    "sample_data": rows[1:6] if len(rows) > 1 else []
                }
        
        return result
    
    if __name__ == "__main__":
        file_path = "planilha_financeira_completa.xlsx"
        data = read_excel(file_path)
        print(json.dumps(data, indent=2, ensure_ascii=False))
        
except ImportError:
    print("ERRO: openpyxl não está instalado. Instale com: pip install openpyxl")
    sys.exit(1)
except Exception as e:
    print(f"ERRO ao ler arquivo: {e}")
    sys.exit(1)

